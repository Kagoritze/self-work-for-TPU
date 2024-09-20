'''
На основе самостоятельной работы №3 реализовать программу, так, чтобы в файл формата *.xlsx, по столбцам, 
записывались все отсортированные массивы с указанием, в начальных ячейках, времени их сортировки и названия алгоритма сортировки, 
а в самый первый столбец файла записать неотсортированный массив. При реализации программы все алгоритмы сортировки должны быть 
представлены в виде функций, а также реализовать функцию записи в файл формата *.xlsx.
'''
import os
import sys
import random
from openpyxl import Workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__),
                                             '../')))

from self_work3.self_work3 import selection_sort, inserts_sort, bubble_sort, shell_sort, quick_sort, measure_time


def write_to_excel(data):
    workbook = Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="Неотсортированный массив")
    for i, value in enumerate(data['original']):
        sheet.cell(row=i + 2, column=1, value=value)

    col = 2
    for sort_name, result in data['sorted'].items():
        header = f"{sort_name}_{result['time']:.6f} сек"
        sheet.cell(row=1, column=col, value=header)

        for i, value in enumerate(result['array']):
            sheet.cell(row=i + 2, column=col, value=value)

        col += 1

    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "sorted_arrays.xlsx")

    workbook.save(file_path)
    print(f"Файл {file_path} успешно сохранен.")


def main():
    arr = [random.uniform(0, 1000) for i in range(1000)]

    data = {'original': arr, 'sorted': {}}

    algorithms = {
        "Выбором": selection_sort,
        "Вставками": inserts_sort,
        "Пузырьком": bubble_sort,
        "Шелла": shell_sort,
        "Быстрая": quick_sort
    }

    for name, sort_func in algorithms.items():
        sorted_arr, elapsed_time = measure_time(sort_func, arr.copy())
        data['sorted'][name] = {'array': sorted_arr, 'time': elapsed_time}

    write_to_excel(data)


if __name__ == "__main__":
    main()
